# coding:utf-8
import os
import sys
import openpyxl
import codecs
import json
import csv
import debug as debug

class xlsx_book(object):

    log = debug.logger("xlsx_book")

    def __init__(self,xlsx,data_only=True):
        xlsx_book.log.enter("__init__(file=" + xlsx + ")")
        self._file = xlsx
        self._data_only = data_only
        self._book = self.open(xlsx,data_only)
        self._dic = {}
        xlsx_book.log.leave()

    def exist(self,xlsx):
        if xlsx == None or xlsx == "": return False
        fname, ext = os.path.splitext(xlsx)        
        if ext != '.xlsx': return False
        if not os.path.isfile(xlsx): return False
        return True
        
    def open(self,xlsx,data_only):
        if not self.exist(xlsx): return None
        return openpyxl.load_workbook(xlsx,data_only=data_only)

    def exist_sheet(self,sheet):
        if not self._book: return False
        if sheet == None or sheet == "": return False
        sheets = self._book.sheetnames
        if sheet in sheets: return True
        return False
    
    def get_sheet_by_name(self,sheet):
        if self.exist_sheet(sheet): return self._book[sheet]
        return None

    def remove_sheet(self, sheet_name):
        if self._data_only: return
        sheet = self.get_sheet_by_name(sheet_name)
        if sheet: _sheet = self._book.remove(sheet)

    def remove_columns(self, sheet_name, cols):
        if self._data_only: return
        sheet = self.get_sheet_by_name(sheet_name)
        if not sheet: return
        cols.sort(reverse=True)        
        for col in cols: sheet.delete_cols(col)            

    def remove_rows(self, sheet_name, rows):
        if self._data_only: return
        sheet = self.get_sheet_by_name(sheet_name)
        if not sheet: return
        rows.sort(reverse=True)                
        for row in rows: sheet.delete_rows(row)
  
    def create_sheet(self, sheet_name, index=0):
        if self._data_only: return
        if self.exist_sheet(sheet_name): return
        self._book.create_sheet(index=index,title=sheet_name)

    def save(self, xlsx=None):
        if self._data_only: return
        if not xlsx: xlsx = self._file
        self._book.save(xlsx)

    def make_json(self, json_path):
        xlsx_book.log.enter("make_json")
        xlsx_book.log.put(name="json_path",value=json_path)        
        json_file = codecs.open(json_path, 'w', 'utf-8')        
        json.dump(self._dic, json_file, indent=2, ensure_ascii=False)
        xlsx_book.log.leave()

    def rows_to_csv(self, rows, csv_path):
        if not isinstance(rows,list): return ""    
        xlsx_book.log.enter("make_csv")
        xlsx_book.log.put(name="csv_path",value=csv_path)
        with open(csv_path, 'w', newline='') as f:        
            writer = csv.writer(f,lineterminator='\n')
            for row in rows:
                if isinstance(row,list): writer.writerow(row)
                else: writer.writerow([row])
        xlsx_book.log.leave()

    def sheet_to_values(self,sheet_name,keys_row=1,id_column=1):
        xlsx_book.log.enter("sheet_to_values")
        sheet = self.get_sheet_by_name(sheet_name)
        if not sheet: xlsx_book.log.leave(); return
        vals = xlsx_values()
        row_num = 0
        for row in sheet.iter_rows():
            row_num += 1
            array = xlsx_book.cells_to_array(row)
            if row_num < keys_row: continue
            if row_num == keys_row:
                vals.set_keys(array,id_column-1)
                continue
            vals.add_values(array)
        xlsx_book.log.leave()
        return vals

    @classmethod
    def cells_to_array(cls,cells):
        array = []
        for cell in cells: array.append(cell.value)
        return array
    
    @classmethod    
    def opt_xlsx(cls, args):
        for arg in args:
            fname, ext = os.path.splitext(arg)        
            if ext == '.xlsx': return arg
        return None

    @classmethod
    def check_xlsx(cls, file):
        if not file: return False
        fname, ext = os.path.splitext(file)
        if ext == '.xlsx': return True
        return False


class xlsx_values(object):

    log = debug.logger("xlsx_values")

    def __init__(self):
        self._id_key = None
        self._keys = []
        self._ids = []
        self._dic = {}

    def set_keys(self, keys, id_index=1):
        self._keys = keys
        self._id_key = keys[id_index]

    def keys(self, skip_blanks=None):
        if skip_blanks: return list(filter(None, self._keys))
        return self._keys

    def ids(self): return self._ids

    def values(self, _id=None):
        if not _id: return self._dic
        return self._dic[_id]

    def sub_values(self, evaluator):
        sub = xlsx_values()
        sub._id_key = self._id_key
        sub._keys = self._keys
        sub._ids = self._ids
        sub._dic = {}
        for key, val in self._dic.items():
            if not evaluator.evaluate(val): continue
            sub._dic[key] = val
        return sub

    def values_by_key(self, key):
        if key not in self._keys: return []
        array = []
        for vals in self._dic.values():
            val = vals[key]
            if not val: continue
            if val not in array: array.append(val)
        return array

    def add_values(self, array):
        _id = None
        vals = {}
        for index in range(0, len(self._keys)):
            key = self._keys[index]
            if not key: continue
            if key == self._id_key: _id = array[index]
            val = array[index]
            if val: vals[key] = val
        if not _id: return
        self._ids.append(_id)
        self._dic[_id] = vals

    def sum(self, key):
        result = 0
        for val in self._dic.values():
            data = val[key]
            if not data: continue
            if isinstance(data, (int, float)): result += data
        return result

    def sumif(self, key, if_key, if_value):
        result = 0
        for val in self._dic.values():
            data = val[key]
            if not data: continue
            if val[if_key] != if_value: continue
            if isinstance(data, (int, float)): result += data
        return result

    def sumifs(self, key, evaluater):
        result = 0
        for val in self._dic.values():
            data = val[key]
            if not data: continue
            if not evaluater.evaluate(val): continue
            if isinstance(data, (int, float)): result += data
        return result

    def debug_print(self):
        xlsx_values.log.put(self._keys, name="keys")
        for key, val in self._dic.items():
            xlsx_values.log.put(key, name="id")
            xlsx_values.log.put(val)

class xlsx_evaluator(object):

    operators = ['==','>','>=','<','<=','!=']

    def __init__(self,key,operator,reference):
        self._key = None
        self._operator = None
        self._ref = None
        if not operator in xlsx_evaluator.operators: return
        self._key = key
        self._operator = operator
        self._ref = reference

    def evaluate(self,dic):
        if not self._key: return False
        if not isinstance(dic,dict): return False
        if not dic[self._key]: return False
        if isinstance(self._ref,list):
            for ref in self._ref:
                if xlsx_evaluator.compair(dic[self._key],self._operator,ref): return True
        return xlsx_evaluator.compair(dic[self._key],self._operator,self._ref)

    @classmethod
    def compair(self,value,operator,reference):
        if not value: return False
        if operator == '==': return value == reference
        if operator == '!=': return value != reference
        if operator == '>': return value > reference
        if operator == '>=': return value >= reference
        if operator == '<': return value < reference
        if operator == '<=': return value <= reference
        return False

class xlsx_evaluators(object):

    def __init__(self,evaluators,and_operator=True):
        self._and = and_operator
        self._evaluators = []
        self.append(evaluators)

    def append(self,evaluators):
        if isinstance(evaluators, xlsx_evaluator): self._evaluators.append(evaluators)
        if isinstance(evaluators,list):
            for evaluator in evaluators:
                if isinstance(evaluator,xlsx_evaluator): self._evaluators.append(evaluator)

    def evaluate(self, dic):
        for evaluator in self._evaluators:
            judge = evaluator.evaluate(dic)
            if self._and == True and judge == False: return False
            if self._and == False and judge == True: return True
        return  self._and


if __name__ == '__main__':
    #debug.start("xlsx_book.debug")    
    args = sys.argv
    xlsx = xlsx_book.opt_xlsx(args)
    if not xlsx: exit
    book = xlsx_book(xlsx,False)
    vals = book.sheet_to_values('detail2',keys_row=2,id_column=2)
    #debug.end()
    

